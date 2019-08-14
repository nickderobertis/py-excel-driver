from openpyxl import load_workbook
import pandas as pd


def load_workbook_active_sheet_into_df(filepath):
    wb, ws = _load_workbook_and_worksheet(filepath)
    return pd.DataFrame(ws.values)

def _load_workbook_and_worksheet(filepath):
    wb = load_workbook(filepath)
    ws = wb.active

    return wb, ws